import cv2,sys,os,time
import tkinter as tk
import numpy as np
from openpyxl import load_workbook,Workbook
from tkinter import ttk
from datetime import date
import json
oldiD=0
names={}
enrol={}
haar_file = face_classifier=r'/home/keshav/my_ten/venv/lib/python3.7/site-packages/cv2/data/haarcascade_frontalface_default.xml'
face_cascade = cv2.CascadeClassifier(haar_file)
width,height = (130,100)
face_recognizer = cv2.face.LBPHFaceRecognizer_create()
def opener(name):
    name="model_files/{}".format(name)
    try:
        f = open(name)
        D=json.load(f)
    except Exception as E:
        print(E)
        D={}
    return D

face_recognizer.read('model_files/model.XML')
model = face_recognizer
names = opener('names.json')
enrol = opener('enrol.json')
names = {int(k):v for k,v in names.items()}
enrol = {int(k):v for k,v in enrol.items()}

def submit2(E1):
    global oldiD,names,model
    name = E1.get()
    newName.destroy()
    save_img(name)
    oldiD,newNames,model=train(oldiD)
    names.update(newNames)
def getname():
    newName = tk.Tk()
    newName.title('Select Section')
    newName.configure(bg='black')
    bigfont = ("gothic",32)
    newName.option_add("*Font", bigfont)
    Label = tk.Label(newName,text='Enter Your Name',bg='black',fg='white').grid(column=1,row=0)
    E1 = tk.Entry(newName,bd =5,bg='black',fg='white')
    E1.grid(column=1,row=1)
    sub=tk.Button(newName,text = 'Submit', command = lambda:submit2(E1,newName)).grid(column=1,row=2)

    newName.mainloop()
def run(path):
    wb = load_workbook(path)
    sheet = wb.active
    b=sheet.max_column
    cell = "{}1"
    print(b)
    ToDa = date.today()
    ToDa = ToDa.strftime("%Y,%m,%d").replace(',','-')
    if b==1:
        sheet["A1"].value = "Enrollment Number"
        sheet["B1"].value = "Name"
        sheet["C1"].value = ToDa
        b=3
    elif sheet[cell.format(chr(b+64))].value != ToDa:
        b+=1
        sheet[cell.format(chr(b+64))].value = ToDa
    b+=64
    wb.save(path)
    trgt="{}{}"
    webcam = cv2.VideoCapture(0)
    flg = False
    count_V = 0
    prev_p = -1
    while True:
        (_,im) = webcam.read()
        gray = cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
        faces = face_cascade.detectMultiScale(gray,1.3,4)
        for (x,y,w,h) in faces:
            cv2.rectangle(im,(x,y),(x+w,y+h),(255,0,0),2)
            face = gray[y:y+h,x:x+w]
            face_resize = cv2.resize(face,(width,height))
            prediction = model.predict(face_resize)
            cv2.rectangle(im,(x,y),(x+w,y+h),(0,255,0),3)
            if prediction[1]<500:
                if prediction[0] == prev_p:
                    count_V+=1
                else:
                    count_V = 0
                prev_p = prediction[0]
                #cv2.rectangle(im, (x-200, y-10), (x + 200, y - 9), (0,0,0), -1)
                cv2.putText(im,'{},{}'.format(names[prediction[0]],prediction[1]//1),(x-200,y-10),cv2.FONT_HERSHEY_PLAIN,2,(255,255,255))
                #print(enrol[prediction[0]])
                if count_V>30:
                    wb = load_workbook(path)
                    sheet = wb.active
                    flg=False
                    i=1
                    for i in range(2,sheet.max_row+1):
                        z="A{}".format(i)
                        #print(z)
                        if sheet[z].value == enrol[prediction[0]]:
                            try:
                                
                                #b=sheet.max_column+1
                                #z="{}{}".format(chr(64+b),i)
                                Ntrgt=trgt.format(chr(b),i)
                                sheet[Ntrgt].value='P'
                                flg=True
                                break
                            except:
                                
                                print(z,chr(b+64),b)
                    if(not flg):
                        try:
                            i+=1
                            #print(i)
                            z="A{}".format(i)
                            sheet[z] = enrol[prediction[0]]
                            z="B{}".format(i)
                            sheet[z] = names[prediction[0]]
                            #b=sheet.max_column+1
                            #z="{}{}".format(chr(b+64),i)
                            Ntrgt=trgt.format(chr(b),i)
                            sheet[Ntrgt]='P'
                        except:
                            print('error',b,chr(b),Ntrgt,"done")
                    wb.save(path)
            else:
                cv2.putText(im,'not recognised',(x-10,y-10),cv2.FONT_HERSHEY_PLAIN,1,(0,255,0))

        cv2.imshow('Face',im)
        key = cv2.waitKey(10)
        if key == 27:
            break
    webcam.release()
    cv2.destroyAllWindows()
def submit1(E1):
    print(E1.get(),"added")
    WB = Workbook()
    ws = WB.active
    name = "{}.xlsx".format(E1.get())
    path = "Sections/{}".format(name)
    WB.save(path)
def submit():
    sc = n.get()
    if sc == 'Add Sections':
        newWindow = tk.Tk()
        newWindow.title('Select Section')
        newWindow.configure(bg='black')
        bigfont = ("gothic",32)
        newWindow.option_add("*Font", bigfont)
        Label = tk.Label(newWindow,text='Enter Section',bg='black',fg='white').grid(column=1,row=0)
        E1 = tk.Entry(newWindow,bd =5,bg='black',fg='white')
        E1.grid(column=1,row=1)
        sub=tk.Button(newWindow,text = 'Submit', command = lambda:submit1(E1)).grid(column=1,row=2)
        newWindow.mainloop()
    else:
        window.destroy()
        pathtoOpen = 'Sections/{}.xlsx'.format(sc)
        run(pathtoOpen)
        
        
def pocom():
    vals = [x[:-5] for x in os.listdir('Sections')]
    vals.sort()
    vals.append('Add Sections')
    SCh['values'] = tuple(vals)
    
window = tk.Tk()
window.title('Select Section')
window.configure(bg='black')
bigfont = ("gothic",32)
window.option_add("*Font", bigfont)
Label = tk.Label(window,text='Selct Section',bg='black',fg='white').grid(column=1,row=0)
n = tk.StringVar()
SCh = ttk.Combobox(window,width=27,textvariable=n,postcommand=pocom,font=("gothic",44))
vals = [x[:-5] for x in os.listdir('Sections')]
vals.sort()
vals.append('Add Sections')
SCh['values'] = tuple(vals)
SCh.grid(column=1,row=1)
SCh.current(0)  
sub_btn=tk.Button(window,text = 'Submit', command = submit) 
sub_btn.grid(column=1,row=3)
window.mainloop()
